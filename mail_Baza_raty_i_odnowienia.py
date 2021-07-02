
import os
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
# import datetime
import time
import re
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from l_h_mailing import passw


start_time = time.time()
os.chdir('/home/robb/Desktop/PROJEKTY/mailing')
now = datetime.now().strftime("Dnia %d.%m.%Y godzina %H:%M:%S")
print('\n' * 3 + '-' * 43 + '\n' + 'MAILING - Przypomnienia o ratach - szuka' + '.' * 3 + f'\n{now}')


class MailingRaty:

    def __init__(self, TEXT):
        self.text = TEXT
        # self.wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
        self.wb = load_workbook(filename="/run/user/1000/gvfs/smb-share:server=192.168.1.12,share=e"
                                         "/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
        self.ws = self.wb['BAZA 2014']
        self.cells = self.ws['T4178':f'BA{self.ws.max_row}']
        today = date.today()
        self.week_period = today - timedelta(-5)  # -5

    def read_excel(self):
        for email, j1, j2, marka, model, nr_rej, rok_prod, SU, j3, j4, j5, pocz, j7, j8, j9, j10, j11, j12, tu, \
            rodz_ub, nr_polisy, j16, j17, j18, j19, j20, j21, j22, przypis, data_raty, kwota, j26, j27, nr_raty \
                in self.cells:
            self.data_raty = data_raty.value
            self.kwota = kwota.value
            self.nr_raty = nr_raty
            self.marka = marka.value
            if self.marka is None:
                self.marka = ''
            self.model = model.value
            if self.model is None:
                self.model = ''
            self.nr_rej = nr_rej.value
            if self.nr_rej is None:
                self.nr_rej = ''
            self.rok_prod = rok_prod.value
            if self.rok_prod is None:
                self.rok_prod = ''
            self.tu = tu.value
            self.rodz_ub = rodz_ub.value
            self.nr_polisy = nr_polisy.value
            self.email = email.value

            yield self.data_raty

    def select_cells(self):
        for self.data_raty in self.read_excel():
            if self.data_raty is not None and re.search('[0-9]', str(self.data_raty)) and not \
                    re.search('[AWV()=.]', str(self.data_raty)):
                data_r = str(self.data_raty)
                self.termin_płatności = data_r[:10]
                if datetime.strptime(str(self.termin_płatności), '%Y-%m-%d').date() == self.week_period and \
                        int(self.nr_raty.value) > 1:
                    if self.email is not None and self.rodz_ub != 'życ':
                        di = {'ALL': 'Allianz', 'AXA': 'AXA', 'COM': 'Compensa', 'EPZU': 'PZU', 'GEN': 'Generali',
                              'GOT': 'Gothaer', 'HDI': 'HDI', 'HES': 'Ergo Hestia', 'IGS': 'IGS', 'INT': 'INTER',
                              'LIN': 'LINK 4', 'MTU': 'MTU', 'PRO': 'Proama', 'PZU': 'PZU', 'RIS': 'InterRisk',
                              'TUW': 'TUW', 'TUZ': 'TUZ', 'UNI': 'Uniqa', 'WAR': 'Warta', 'WIE': 'Wiener',
                              'YCD': 'You Can Drive'}
                        self.tu = di.get(self.tu)

                        yield self.termin_płatności

    def iterate_funct(self):
        for _ in self.select_cells():
            print()
            print('Data raty: ' + str(self.termin_płatności))
            print('Marka/kod poczt: ' + str(self.marka))
            print('Model/miasto: ' + str(self.model))
            print('Nr.rej./ulica: ' + str(self.nr_rej))
            print('Rok prod.: ' + str(self.rok_prod))
            print('TU: ' + str(self.tu))
            print('Nr. polisy: ' + str(self.nr_polisy))
            print('Kwota: ' + str(self.kwota))
            print(self.email)

            dash = " - "
            comma_year = ", rok "
            if self.marka == '' and self.model == '':
                dash = ""
                comma_year = ""

            if self.rok_prod == '':
                comma_year = ''

            SUBJECT = "Przypomnienie o płatności raty."
            text_html = f"""
                        {self.text}
                    {SUBJECT}
                 </td>
               </tr>
               <tr>
                 <td class="free-text"><br>
                    Dnia {str(self.termin_płatności)} upływa termin wpłaty raty za polisę.<br>
                    Polisa nr.: {str(self.nr_polisy)} - T.U. {self.tu}<br> 
                    Kwota: {str(self.kwota)} zł<br> 
                    {str(self.marka)} {str(self.model)} {dash} {str(self.nr_rej)}{comma_year} {str(self.rok_prod)}<br>
                    <br><br>Prosimy o terminową wpłatę,<br>
                    <a href="ubezpieczenia-magro.pl">www.ubezpieczenia-magro.pl</a><br><br>
                    Zapraszamy do zapoznania sie z naszą ofertą ubezpieczeń na życie.<br>
                    Szczegóły w załącznikch.
               </td>
             </tr>
             <tr>
               <td class="mini-block-container">
                 <table cellspacing="0" cellpadding="0" width="100%"  style="border-collapse:separate !important;">
                   <tr>
        
                     <td class="m_-8810930916023015007user-msg user-msg">
            """

            html = text_html
            text_alt = f"""
                    Przypomnienie o płatności raty.\n
                    Dnia {str(self.termin_płatności)} upływa termin wpłaty raty za polisę.\n
                    Polisa nr.:  {str(self.nr_polisy)} - T.U. {self.tu}
                    Kwota: {str(self.kwota)} zł
                    {str(self.marka)} {str(self.model)} {dash} {str(self.nr_rej)}{comma_year} {str(self.rok_prod)}\n
                    Prosimy o terminową wpłatę,
                    ubezpieczenia-magro.pl
                    MAGRO Ubezpieczenia Sp. z o.o.
                    Zapraszamy do zapoznania sie z naszą ofertą ubezpieczeń na życie.\n
                    Szczegóły w załącznikch.
                    """

            mail = MIMEMultipart('alternative')
            mail['Subject'] = 'MAGRO Ubezpieczenia Sp. z o.o.'
            mail['From'] = 'przypomnienia@ubezpieczenia-magro.pl'
            mail['To'] = self.email  # Do prob zmienic email
            mail['Cc'] = 'ubezpieczenia.magro@gmail.com'

            ulotka_gen = 'Generali_GO_PLUS.pdf'
            ulotka_war = 'Warta_WDCIR.pdf'
            for attachment in ulotka_gen, ulotka_war:
                my_file = MIMEBase('application', 'pdf')
                with open(attachment, 'rb') as f:
                    my_file.set_payload(f.read())
                    my_file.add_header('Content-Disposition', f'attachment; filename = {attachment}', )
                    encoders.encode_base64(my_file)
                    mail.attach(my_file)

            part1 = MIMEText(text_alt, 'plain')
            part2 = MIMEText(html, 'html')
            mail.attach(part1)
            mail.attach(part2)
            msg_full = mail.as_string().encode('utf-8')
            server = smtplib.SMTP('ubezpieczenia-magro.pl', 587)
            server.starttls()
            server.login('przypomnienia@ubezpieczenia-magro.pl', passw)
            server.sendmail('przypomnienia@ubezpieczenia-magro.pl', [self.email, 'ubezpieczenia.magro@gmail.com'], msg_full)
            # server.sendmail('przypomnienia@ubezpieczenia-magro.pl', ['magro@ubezpieczenia-magro.pl'], msg_full)  # Do prob
            server.quit()

            if self.email is not None:
                print('\nWysłane przypomnienia o nadchodzących ratach!')
            else:
                print('Brak adresu email')


TEXT = """
             <!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
         <html xmlns="http://www.w3.org/1999/xhtml">
         <head>
           <meta http-equiv="Content-Type" content="text/html; charset=UTF-8" />
           <meta name="viewport" content="width=device-width, initial-scale=1" />
           <title>Oxygen Invitation</title>

           <style type="text/css">
             /* Take care of image borders and formatting, client hacks */
             img { max-width: 600px; outline: none; text-decoration: none; -ms-interpolation-mode: bicubic;}
             a img { border: none; }
             table { border-collapse: collapse !important;}
             #outlook a { padding:0; }
             .ReadMsgBody { width: 100%; }
             .ExternalClass { width: 100%; }
             .backgroundTable { margin: 0 auto; padding: 0; width: 100% !important; }
             table td { border-collapse: collapse; }
             .ExternalClass * { line-height: 115%; }
             .container-for-gmail-android { min-width: 600px; }


             /* General styling */
             * {
               font-family: Helvetica, Arial, sans-serif;
             }

             body {
               -webkit-font-smoothing: antialiased;
               -webkit-text-size-adjust: none;
               width: 100% !important;
               margin: 0 !important;
               height: 100%;
               color: #676767;
             }

             td {
               font-family: Helvetica, Arial, sans-serif;
               font-size: 14px;
               color: #777777;
               text-align: center;
               line-height: 21px;
             }

             a {
               color: #fe5100; font-size:1.6em;
               text-decoration: none !important;
             }
             .kla{color:#fff !important;}

             .pull-left {
               text-align: left;
             }

             .pull-right {
               text-align: right;
             }

             .header-lg,
             .header-md,
             .header-sm {
               font-size: 25px;
               font-weight: 700;
               line-height: normal;
               padding: 35px 0 0;
               color: #4d4d4d;
             }

             .header-md {
               font-size: 24px;
             }

             .header-sm {
               padding: 5px 0;
               font-size: 18px;
               line-height: 1.3;
             }

             .content-padding {
               padding: 20px 0 30px;
             }

             .mobile-header-padding-right {
               width: 20%;
               text-align: right;
               padding-left: 10px;
             }

             .mobile-header-padding-left {
               width: 80%;
               text-align: left;
               padding-left: 10px;
             }

             .free-text {
               width: 100% !important;
               padding: 10px 60px 0px;
             }

             .block-rounded {
               border-radius: 5px;
               border: 1px solid #e5e5e5;
               vertical-align: top;
             }

             .button {
               padding: 30px 0 0;
             }

             .info-block {
               padding: 0 20px;
               width: 260px;
             }

             .mini-block-container {
               padding: 30px 50px;
               width: 500px;
             }

             .mini-block {
             background-color: #ffffff;
             width: 93%;
             border: 1px solid #cccccc;
             border-radius: 5px;
             padding: 34px 10px;
             margin: 15px 0px;
                display: block;

             }

             .block-rounded {
               width: 260px;
             }

             .info-img {
               width: 258px;
               border-radius: 5px 5px 0 0;
             }

             .force-width-img {
               width: 480px;
               height: 1px !important;
             }

             .force-width-full {
               width: 600px;
               height: 1px !important;
             }

             .user-img img {
               width: 130px;
               border-radius: 5px;
               border: 1px solid #cccccc;
             }

             .user-img {
               text-align: center;
               border-radius: 100px;
               color: #ff6f6f;
               font-weight: 700;
             }

             .user-msg {
               padding-top: 10px;
               font-size: 14px;
               text-align: center;
               font-style: italic;
             }

             .mini-img {
               padding: 5px;
               WIDTH:48%;

             }

             .mini-img img {
               border-radius: 5px;
               WIDTH:100%;
              min-height: 100px;
             }

             .force-width-gmail {
               min-width:600px;
               height: 0px !important;
               line-height: 1px !important;
               font-size: 1px !important;
             }

             .mini-imgs {
               padding: 25px 0 30px;
             }
           </style>

           <style type="text/css" media="screen">
             @import url(https://fonts.googleapis.com/css?family=Oxygen:400,700);
           </style>

           <style type="text/css" media="screen">
             @media screen {
               /* Thanks Outlook 2013! */
               * {
                 font-family: 'Oxygen', 'Helvetica Neue', 'Arial', 'sans-serif' !important;
               }
             }
           </style>

           <style type="text/css" media="only screen and (max-width: 480px)">
             /* Mobile styles */
             @media only screen and (max-width: 480px) {

               table[class*="container-for-gmail-android"] {
                 min-width: 290px !important;
                 width: 100% !important;
               }

               table[class="w320"] {
                 width: 320px !important;
               }

               img[class="force-width-gmail"] {
                 display: none !important;
                 width: 0 !important;
                 height: 0 !important;
               }

               td[class*="mobile-header-padding-left"] {
                 width: 80% !important;
                 padding-left: 0 !important;
               }

               td[class*="mobile-header-padding-right"] {
                 width: 20% !important;
                 padding-right: 0 !important;
               }

               td[class="mobile-block"] {
                 display: block !important;
               }

               td[class="mini-img"],
               td[class="mini-img"] img{
                 width: 150px !important;min-height: 100px;
               }

               td[class="header-lg"] {
                 font-size: 24px !important;
                 padding-bottom: 5px !important;
               }

               td[class="header-md"] {
                 font-size: 18px !important;
                 padding-bottom: 5px !important;
               }

               td[class="content-padding"] {
                 padding: 5px 0 30px !important;
               }

               td[class="button"] {
                 padding: 5px !important;
               }

               td[class*="free-text"] {
                 padding: 10px 18px 30px !important;
               }

               img[class="force-width-img"],
               img[class="force-width-full"] {
                 display: none !important;
               }

               td[class="info-block"] {
                 display: block !important;
                 width: 280px !important;
                 padding-bottom: 40px !important;
               }

               td[class="info-img"],
               img[class="info-img"] {
                 width: 278px !important;
               }

               td[class="mini-block-container"] {
                 padding: 8px 20px !important;
                 width: 280px !important;
               }

               td[class="mini-block"] {
                 padding: 20px !important;
               }

               td[class="user-img"] {
                 display: block !important;
                 text-align: center !important;
                 width: 100% !important;
                 padding-bottom: 10px;
               }

               td[class="user-msg"] {
                 display: block !important;
                 padding-bottom: 20px;
               }
             }

            td.m_-8810930916023015007user-msg a span {
             float: left;
             font-size: 28px;
             color: #555 !important;
             font-weight: bold;
             margin: 12px;
             margin: -10px 12px 0px;
         }

         td.m_-8810930916023015007user-msg i {
             float: right;
             font-size: 17px;
             color: #555 !important;
             font-weight: 600;
             color: #fe5000 !important;
             margin: -10px 0px;
         }

         .link{
         background-color:#ff6f6f;border-radius:5px;color:#ffffff;display:inline-block;font-family:'Cabin',Helvetica,Arial,sans-serif;font-size:14px;font-weight:regular;line-height:45px;text-align:center;text-decoration:none;width:155px;
         }
         .logo{
         color: #4CAF50;
             font-size: 35px;
             font-weight:bold;

             color: #4d4d4d;
         }
           </style>
         </head>

         <body bgcolor="#f7f7f7">
         <table align="center" cellpadding="0" cellspacing="0" class="container-for-gmail-android" width="100%">
           <tr>
             <td align="left" valign="top" width="100%" style="background:repeat-x url(http://s3.amazonaws.com/swu-filepicker/4E687TRe69Ld95IDWyEg_bg_top_02.jpg) #ffffff;">
               <center>
               <img src="https://s3.amazonaws.com/swu-filepicker/SBb2fQPrQ5ezxmqUTgCr_transparent.png" class="force-width-gmail">
                 <table cellspacing="0" cellpadding="0" width="100%" bgcolor="#ffffff" background="https://s3.amazonaws.com/swu-filepicker/4E687TRe69Ld95IDWyEg_bg_top_02.jpg" style="background-color:transparent">
                   <tr>
                     <td width="100%" height="80" valign="top" style="text-align: center; vertical-align:middle;">
                     <!--[if gte mso 9]>
                     <v:rect xmlns:v="urn:schemas-microsoft-com:vml" fill="true" stroke="false" style="mso-width-percent:1000;height:80px; v-text-anchor:middle;">
                       <v:fill type="tile" src="http://s3.amazonaws.com/swu-filepicker/4E687TRe69Ld95IDWyEg_bg_top_02.jpg" color="#ffffff" />
                       <v:textbox inset="0,0,0,0">
                     <![endif]-->
                       <center>
                         <table cellpadding="0" cellspacing="0" width="600" class="w320">
                           <tr>
                             <td class="pull-left mobile-header-padding-left" style="vertical-align: middle;">
                               <a href="" class="logo">MAGRO<!--<img width="137" height="47" src="http://s3.amazonaws.com/swu-filepicker/0zxBZVuORSxdc9ZCqotL_logo_03.gif" alt="logo">--></a>
                             </td>
                             <td class="pull-right mobile-header-padding-right" style="color: #4d4d4d;">

                               <a href="https://www.facebook.com/ubezpieczeniaMagro/"><img width="38" height="47" src="https://s3.amazonaws.com/swu-filepicker/LMPMj7JSRoCWypAvzaN3_social_09.gif" alt="facebook" /></a>

                             </td>
                           </tr>
                         </table>
                       </center>
                       <!--[if gte mso 9]>
                       </v:textbox>
                     </v:rect>
                     <![endif]-->
                     </td>
                   </tr>
                 </table>
               </center>
             </td>
           </tr>
           <tr>
             <td align="center" valign="top" width="100%" style="background-color: #f7f7f7;" class="content-padding">
               <center>
                 <table cellspacing="0" cellpadding="0" width="600" class="w320">
                   <tr>
                     <td class="header-lg">"""


# raty = MailingRaty(TEXT)
# raty.read_excel()
# raty.select_cells()
# raty.iterate_funct()

end_time = time.time() - start_time
print('\nCzas wykonania: {:.0f} sekund'.format(end_time))

########################################################################################################################

start_time = time.time()
now = datetime.now().strftime("Dnia %d.%m.%Y godzina %H:%M:%S")
print('\n' * 3 + 'MAILING - Przypomnienia o wznowieniach - szuka' + '.' * 3 + f'\n{now}')


class MailingOdn:

    def __init__(self, TEXT):
        self.text = TEXT
        # self.wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
        self.wb = load_workbook(filename="/run/user/1000/gvfs/smb-share:server=192.168.1.12,share=e/Agent baza/"
                                         "2014 BAZA MAGRO.xlsx", read_only=True)
        self.ws = self.wb['BAZA 2014']
        self.cells = self.ws['G4178':f'BA{self.ws.max_row}']
        today = date.today()
        self.week_period_odn = today - timedelta(-14)

    @staticmethod
    def remove_html_tags(txt):
        clean = re.compile('<.*?>')
        return re.sub(clean, '', txt)

    def read_excel(self):
        for rozlicz, H, I, J, K, L, M, N, O, P, Q, R, nr_tel, email, U, V, marka, model, przedmiot_ub, rok_prod, SU, \
                AB, AC, AD, pocz, koniec, AG, AH, AI, AJ, AK, tu, rodz_ub, nr_polisy, AO, AP, AQ, AR, AS, AT, AU, \
                    przypis, data_raty, kwota, AY, AZ, nr_raty in self.cells:
            self.rozlicz = rozlicz.value
            self.email = email.value
            self.marka = marka.value
            if self.marka is None:
                self.marka = ''
            self.model = model.value
            if self.model is None:
                self.model = ''
            self.przedmiot_ub = przedmiot_ub.value
            if self.przedmiot_ub is None:
                self.przedmiot_ub = ''
            self.rok_prod = rok_prod.value
            if self.rok_prod is None:
                self.rok_prod = ''
            self.koniec = koniec.value
            self.tu = tu.value
            self.rodz_ub = rodz_ub.value
            self.nr_polisy = nr_polisy.value
            self.przypis = przypis.value
            self.data_raty = data_raty.value
            self.kwota = kwota.value
            self.nr_raty = nr_raty

            yield self.koniec


    def select_cells_odn(self):
        for self.koniec in self.read_excel():
            if self.koniec is not None and re.search('[0-9]', str(self.koniec)):
                # and not \
                #     re.search('[AWV()=.]', str(self.koniec)):

                koniec_okresu = str(self.koniec)
                self.koniec_okresu_bez_sec = koniec_okresu[:10]
                if datetime.strptime(str(self.koniec_okresu_bez_sec), '%Y-%m-%d').date() == self.week_period_odn:
                    if self.email is not None and self.rodz_ub != 'życ' and self.przypis is not None:

                        d = {'Filipiak': 'Ultimatum, tel. 694888197',
                             'Wawrzyniak': 'A. Wawrzyniak, tel. 691602675',
                             'Wołowski': 'M. Wołowskim, tel. 692830084',
                             'Robert': 'naszym biurem, tel. 572 810 576<br>\n'
                             '<a href="ubezpieczenia.magro@gmail.com" style="font-size: 16px">'
                             'ubezpieczenia.magro@gmail.com</a>',
                             }
                        if self.rozlicz in d:
                            self.rozlicz = d.get(self.rozlicz)
                        else:
                            self.rozlicz = 'naszym biurem,<br>' \
                                           'tel. 602 752 893 lub 42 637 19 97<br>\n' \
                                           '<a href="ubezpieczenia.magro@gmail.com" style="font-size: 16px">' \
                                           'ubezpieczenia.magro@gmail.com</a>'

                        di = {'ALL': 'Allianz', 'AXA': 'AXA', 'COM': 'Compensa', 'EPZU': 'PZU', 'GEN': 'Generali',
                              'GOT': 'Gothaer', 'HDI': 'HDI', 'HES': 'Ergo Hestia', 'IGS': 'IGS', 'INT': 'INTER',
                              'LIN': 'LINK 4', 'MTU': 'MTU', 'PRO': 'Proama', 'PZU': 'PZU', 'RIS': 'InterRisk',
                              'TUW': 'TUW', 'TUZ': 'TUZ', 'UNI': 'Uniqa', 'WAR': 'Warta', 'WIE': 'Wiener',
                              'YCD': 'You Can Drive'
                              }
                        self.tu = di.get(self.tu)

                        yield self.koniec_okresu_bez_sec


    def iterate_funct_odn(self):
        for _ in self.select_cells_odn():
            print()
            print('Data końca polisy: ' + self.koniec_okresu_bez_sec)
            print('Marka/kod poczt: ' + str(self.marka))
            print('Model/miasto: ' + str(self.model))
            print('Nr.rej./ulica: ' + str(self.przedmiot_ub))
            print('Rok prod.: ' + str(self.rok_prod))
            print('TU: ' + str(self.tu))
            print('Nr. polisy: ' + str(self.nr_polisy))
            print('Przypis z polisy: ' + str(self.przypis))
            print(self.email)


            dash = " - "
            comma_year = ", rok "
            if self.marka == '' and self.model == '':
                dash = ""
                comma_year = ""

            if self.rok_prod == '':
                comma_year = ''

            SUBJECT = "Przypomnienie o końcu ochrony."
            TEXT = f"""{self.text}
                    {SUBJECT}
             </td>
           </tr>
           <tr>
             <td class="free-text">
            <br>
                Dnia {str(self.koniec_okresu_bez_sec)} dobiega końca Twoja polisa ubezpieczeniowa.<br>
                Polisa nr.: {str(self.nr_polisy)} - T.U. {self.tu}<br>
                {str(self.marka)} {str(self.model)} {dash} {str(self.przedmiot_ub)}{comma_year} {str(self.rok_prod)}
                <br><br>W sprawie odnowienia prosimy o kontakt<br>z {str(self.rozlicz)}<br><br><br><br>
                <a href="ubezpieczenia-magro.pl/kalkulatorOC">Kalkulator ubezpieczenia OC</a><br><br>
                Zapraszamy do zapoznania sie z naszą ofertą ubezpieczeń na życie.<br>
                Szczegóły w załącznikch.
               </td>
           </td>
         </tr>
         <tr>
           <td class="mini-block-container">
             <table cellspacing="0" cellpadding="0" width="100%"  style="border-collapse:separate !important;">
               <tr>
    
                 <td class="m_-8810930916023015007user-msg user-msg">
    """
            os_do_kontaktu = self.remove_html_tags(str(self.rozlicz))
            html = TEXT
            text_alt = f"""
Przypomnienie o końcu ochrony.\n
Dnia {str(self.koniec_okresu_bez_sec)} dobiega końca Twoja polisa ubezpieczeniowa.\n
Polisa nr.: {str(self.nr_polisy)} - T.U. {self.tu}
{str(self.marka)} {str(self.model)} {dash} {str(self.przedmiot_ub)} {comma_year} {str(self.rok_prod)}\n
W sprawie odnowienia prosimy o kontakt\n
z {os_do_kontaktu}.\n
ubezpieczenia-magro.pl/kalkulatorOC\n
MAGRO Ubezpieczenia Sp. z o.o.\n
Zapraszamy do zapoznania sie z naszą ofertą ubezpieczeń na życie.\n
Szczegóły w załącznikch.
"""

            mail = MIMEMultipart('alternative')
            mail['Subject'] = 'MAGRO Ubezpieczenia Sp. z o.o.'
            mail['From'] = 'przypomnienia@ubezpieczenia-magro.pl'
            mail['To'] = self.email  # Do prob zmienic email
            mail['Cc'] = 'ubezpieczenia.magro@gmail.com'

            ulotka_gen = 'Generali_GO_PLUS.pdf'
            ulotka_war = 'Warta_WDCIR.pdf'
            for attachment in ulotka_gen, ulotka_war:
                my_file = MIMEBase('application', 'pdf')
                with open(attachment, 'rb') as f:
                    my_file.set_payload(f.read())
                    my_file.add_header('Content-Disposition', f'attachment; filename = {attachment}', )
                    encoders.encode_base64(my_file)
                    mail.attach(my_file)

            part1 = MIMEText(text_alt, 'plain')
            part2 = MIMEText(html, 'html')
            mail.attach(part1)
            mail.attach(part2)
            msg_full = mail.as_string().encode('utf-8')
            server = smtplib.SMTP('ubezpieczenia-magro.pl', 587)
            server.starttls()
            server.login('przypomnienia@ubezpieczenia-magro.pl', passw)
            server.sendmail('przypomnienia@ubezpieczenia-magro.pl', [self.email, 'ubezpieczenia.magro@gmail.com'], msg_full)
            # server.sendmail('przypomnienia@ubezpieczenia-magro.pl', ['magro@ubezpieczenia-magro.pl'], msg_full)  # do prob
            server.quit()

            if self.email is not None:
                print('\nWysłane przypomnienia o końcu polis!')
            if self.email == '':
                print('Brak adresu email')


raty = MailingOdn(TEXT)
raty.read_excel()
raty.select_cells_odn()
raty.iterate_funct_odn()

end_time = time.time() - start_time
print('\nCzas wykonania: {:.0f} sekund'.format(end_time))
time.sleep(1)
